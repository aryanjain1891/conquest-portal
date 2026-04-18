"""
Backfill UserProfile rows with richer data from the Conquest '24 Excel.
Runs inside the web container: docker exec conquest_back_web python scripts/ingest_excel.py /home/app/web/portal_info.xlsx
"""
import sys, os, django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conquest_back.settings")
django.setup()

import openpyxl
from users.models import UserProfile

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "/home/app/web/portal_info.xlsx"

# Column index in each sheet → UserProfile field. None means skip.
SHEETS = {
    "Mentors":  {"role_filter": "Mentor",
                 "map": {"Mentor Name": "name", "Designation": "designation",
                         "Organisation": "company_name", "LinkedIn": "linkedin",
                         "Picture": "profile_logo", "One Liner": "description",
                         "Expertise": "sector_of_expertise"}},
    "Coaches":  {"role_filter": "Coach",
                 "map": {"Coach Name": "name", "Designation": "designation",
                         "Organisation": "company_name", "LinkedIn": "linkedin",
                         "Picture": "profile_logo", "Description": "description",
                         "Expertise": "sector_of_expertise", "Location": "location"}},
    "Experts":  {"role_filter": "Function Expert",
                 "map": {"Expert Name": "name", "Designation": "designation",
                         "Organisation": "company_name", "LinkedIn": "linkedin",
                         "Picture": "profile_logo", "One Liner": "description",
                         "Verticals": "verticals", "Business Model": "business_models",
                         "Horizontals": "horizontals",
                         "Functions of Expertise": "function_of_expertise"}},
    "Angels":   {"role_filter": "Angel",
                 "map": {"Name": "name", "Designation": "designation",
                         "Organisation": "company_name", "LinkedIn": "linkedin",
                         "Picture": "profile_logo", "One Liner": "description"}},
}

def norm(s): return (s or "").strip().lower()

def run():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    totals = {"updated": 0, "created": 0, "skipped_no_name": 0, "matched": 0}

    for sheet_name, cfg in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        col_index = {col_name: i for i, col_name in enumerate(header) if col_name}
        role = cfg["role_filter"]

        # Index existing profiles for this role by normalized name
        existing = {norm(p.name): p for p in UserProfile.objects.filter(role__iexact=role) if p.name}

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name:
                totals["skipped_no_name"] += 1
                continue

            profile = existing.get(norm(name))
            if not profile:
                continue  # don't create new users here; only backfill existing
            totals["matched"] += 1

            changed = False
            for excel_col, field in cfg["map"].items():
                if excel_col not in col_index:
                    continue
                val = row[col_index[excel_col]]
                if val is None or (isinstance(val, str) and not val.strip()):
                    continue
                current = getattr(profile, field, None)
                if current and str(current).strip():
                    continue  # non-destructive: keep existing
                setattr(profile, field, str(val).strip())
                changed = True
            if changed:
                profile.save()
                totals["updated"] += 1

        print(f"[{sheet_name}] rows: {len(rows)-1}  matched: {totals['matched']}  updated so far: {totals['updated']}")

    print("\nDone.", totals)

if __name__ == "__main__":
    run()
